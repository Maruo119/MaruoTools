import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

def find_function_declaration(file_path, line_number):
    """
    Searches upwards in a file from a specific line number to find the enclosing function declaration.
    For Java files, it looks for function definitions such as those containing 'public', 'private', 'protected', or 'static'.
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for i in range(line_number - 2, -1, -1):  # Search backwards from the line above the hit
                if re.search(r'\b(public|private|protected|static)\b.*\b([a-zA-Z_][a-zA-Z0-9_]*)\b\s*\(', lines[i]):
                    return lines[i].strip()
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
    return "Unknown Function Declaration"

def search_keyword_in_repository(repo_path, keyword, output_excel):
    results = []

    # Walk through the repository
    for root, dirs, files in os.walk(repo_path):
        for file in files:
            if file.endswith(".java"):  # Process only Java files
                file_path = os.path.join(root, file)

                # Only process text files
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        for line_number, line in enumerate(f, start=1):
                            if re.search(keyword, line):
                                function_declaration = find_function_declaration(file_path, line_number)
                                results.append((file_path, line_number, function_declaration, line.strip()))
                except (UnicodeDecodeError, PermissionError):
                    # Skip non-text files or files without read permissions
                    continue

    # Create Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Search Results"

    # Add headers
    headers = ["File Path", "Line Number", "Function Declaration", "Line Content"]
    for col_num, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Add data
    for row_num, (file_path, line_number, function_declaration, line_content) in enumerate(results, start=2):
        sheet[f"A{row_num}"] = file_path
        sheet[f"B{row_num}"] = line_number
        sheet[f"C{row_num}"] = function_declaration
        sheet[f"D{row_num}"] = line_content

    # Save the Excel file
    workbook.save(output_excel)
    print(f"Results saved to {output_excel}")

if __name__ == "__main__":
    # User inputs
    #repository_path = input("Enter the path to the repository: ").strip()
    #search_keyword = input("Enter the keyword to search for: ").strip()
    #output_excel_path = input("Enter the output Excel file path: ").strip()

    repository_path = "/Users/yuta/Documents/examples-main".strip()
    search_keyword = "draw".strip()
    output_excel_path = "/Users/yuta/backlog_python/search_impact/output_v2.xlsx".strip()



    # Run the search
    search_keyword_in_repository(repository_path, search_keyword, output_excel_path)
